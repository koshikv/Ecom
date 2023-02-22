import openpyxl


def getRowCount(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.max_row


def getCloumnCout(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.max_column


def readData(file, sheetname, rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    readvalues = sheet.cell(row=rownum, column=columnno)
    return readvalues.value


def writeData(file, sheetname, rownum, columnno, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    writevalues = sheet.cell(row=rownum, column=columnno)
    writevalues.value = data
    workbook.save(file)
